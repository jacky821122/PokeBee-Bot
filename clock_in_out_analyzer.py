from __future__ import annotations

import argparse
import csv
import re
from dataclasses import dataclass
from datetime import datetime, timedelta
from pathlib import Path
from typing import Optional

FULL_TIME_NAMES = {"小王叭"}
TIME_FORMAT = "%Y-%m-%d %H:%M:%S"


@dataclass
class Event:
    kind: str
    timestamp: Optional[datetime] = None


@dataclass
class PairRecord:
    employee: str
    date: str
    shift: str
    in_raw: str
    in_norm: str
    out_raw: str
    out_norm: str
    normal_hours: float
    overtime_hours: float
    note: str


@dataclass
class EmployeeSummary:
    employee: str
    is_full_time: bool
    normal_hours: float = 0.0
    overtime_hours: float = 0.0
    specials: list[str] = None

    def __post_init__(self) -> None:
        if self.specials is None:
            self.specials = []


def round_to_half_hour(dt: datetime) -> datetime:
    if dt.minute < 15:
        return dt.replace(minute=0, second=0, microsecond=0)
    if dt.minute < 45:
        return dt.replace(minute=30, second=0, microsecond=0)
    return (dt + timedelta(hours=1)).replace(minute=0, second=0, microsecond=0)


def floor_to_half_hour(dt: datetime) -> datetime:
    return dt.replace(minute=30 if dt.minute >= 30 else 0, second=0, microsecond=0)

def ceiling_to_half_hour(dt: datetime) -> datetime:
    if dt.minute <= 30:
        return dt.replace(minute=30, second=0, microsecond=0)
    else:
        return (dt + timedelta(hours=1)).replace(minute=0, second=0, microsecond=0)

def normalize_in_time(in_ts: datetime) -> datetime:
    d = in_ts.date()
    if in_ts < datetime(d.year, d.month, d.day, 10, 0):
        return ceiling_to_half_hour(in_ts)
    return round_to_half_hour(in_ts)

def normalize_out_time(out_ts: datetime, normal_end: Optional[datetime]) -> datetime:
    if normal_end is None or out_ts <= normal_end:
        return round_to_half_hour(out_ts)
    if out_ts < normal_end + timedelta(minutes=30):
        return normal_end
    return floor_to_half_hour(out_ts)


def fmt_hours(hours: float) -> str:
    if abs(hours - round(hours)) < 1e-9:
        return f"{int(round(hours))}"
    return f"{hours:.1f}".rstrip("0").rstrip(".")


def parse_csv(path: Path) -> dict[str, list[Event]]:
    employees: dict[str, list[Event]] = {}
    current_name: Optional[str] = None
    pending_no_in = False

    with path.open("r", encoding="utf-8-sig", newline="") as f:
        for row in csv.reader(f):
            if not row:
                continue
            c0 = (row[0] or "").strip()
            c1 = (row[1] if len(row) > 1 else "").strip()

            if c0 == "" and c1 == "":
                continue

            if c1 == "" and c0 not in {"clock-in", "clock-out", "no clock-in record", "no clock-out record"} and not c0.startswith("Total hours"):
                current_name = c0
                employees.setdefault(current_name, [])
                pending_no_in = False
                continue

            if current_name is None:
                continue

            if c0.startswith("Total hours"):
                current_name = None
                pending_no_in = False
                continue

            if c0 == "clock-in" and c1:
                employees[current_name].append(Event("clock-in", datetime.strptime(c1, TIME_FORMAT)))
            elif c0 == "clock-out" and c1:
                kind = "clock-out-no-in" if pending_no_in else "clock-out"
                employees[current_name].append(Event(kind, datetime.strptime(c1, TIME_FORMAT)))
                pending_no_in = False
            elif c0 == "no clock-in record":
                pending_no_in = True
            elif c0 == "no clock-out record":
                employees[current_name].append(Event("no-clock-out"))

    return {k: v for k, v in employees.items() if v}


def classify_shift(norm_in: datetime) -> tuple[str, datetime]:
    d = norm_in.date()
    h = norm_in.hour + norm_in.minute / 60
    if h < 14:
        return "早班", datetime(d.year, d.month, d.day, 14, 0)
    elif h <= 16 and h >= 14:
        return "晚班1", datetime(d.year, d.month, d.day, 20, 0)
    elif h > 16:
        return "晚班2", datetime(d.year, d.month, d.day, 20, 30)


def add_record(records: list[PairRecord], summary: EmployeeSummary, rec: PairRecord, force_special: bool = False) -> None:
    records.append(rec)
    summary.normal_hours += rec.normal_hours
    summary.overtime_hours += rec.overtime_hours
    if force_special or rec.note:
        summary.specials.append(f"{rec.date} {rec.note}")


def handle_full_time(summary: EmployeeSummary, records: list[PairRecord], name: str, in_ts: Optional[datetime], out_ts: Optional[datetime], inferred_no_in: bool) -> None:
    date = (in_ts or out_ts).date().isoformat()
    in_norm = normalize_in_time(in_ts) if in_ts else None
    normal_end = datetime(out_ts.year, out_ts.month, out_ts.day, 20, 0) if out_ts else None
    out_norm = normalize_out_time(out_ts, normal_end) if out_ts else None

    normal = 8.0 if (in_ts and out_ts) else 0.0
    overtime = 0.0
    notes: list[str] = []

    if not in_ts or not out_ts or inferred_no_in:
        notes.append("缺打卡，需人工確認")
    elif out_norm >= normal_end + timedelta(minutes=30):
        overtime = (out_norm - normal_end).total_seconds() / 3600
        notes.append(f"下班 {out_norm.strftime('%H:%M')}，計為 {fmt_hours(overtime)} 小時加班")

    add_record(
        records,
        summary,
        PairRecord(
            employee=name,
            date=date,
            shift="正職",
            in_raw=in_ts.strftime(TIME_FORMAT) if in_ts else "",
            in_norm=in_norm.strftime("%Y-%m-%d %H:%M") if in_norm else "",
            out_raw=out_ts.strftime(TIME_FORMAT) if out_ts else "",
            out_norm=out_norm.strftime("%Y-%m-%d %H:%M") if out_norm else "",
            normal_hours=normal,
            overtime_hours=overtime,
            note="；".join(notes),
        ),
        force_special=bool(notes),
    )


def handle_hourly(summary: EmployeeSummary, records: list[PairRecord], name: str, in_ts: Optional[datetime], out_ts: Optional[datetime], inferred_no_in: bool) -> None:
    date = (in_ts or out_ts).date().isoformat()
    in_norm = normalize_in_time(in_ts) if in_ts else None

    shift = "未知"
    normal = 0.0
    overtime = 0.0
    notes: list[str] = []

    out_norm: Optional[datetime] = None
    if in_norm and out_ts:
        shift_preview, normal_end_preview = classify_shift(in_norm)
        _ = shift_preview
        out_norm = normalize_out_time(out_ts, normal_end_preview)
    elif out_ts:
        out_norm = round_to_half_hour(out_ts)

    if in_norm and out_norm and in_norm < in_norm.replace(hour=14, minute=0) and out_norm >= out_norm.replace(hour=20, minute=0):
        shift = "全日連續班"
        normal = 8.0
        late_end = out_norm.replace(hour=20, minute=30)
        if out_norm > late_end:
            overtime = (out_norm - late_end).total_seconds() / 3600
        notes.append("全日連續班（強制拆分）")
    elif in_norm and out_norm:
        shift, _ = classify_shift(in_norm)
        worked_hours = (out_norm - in_norm).total_seconds() / 3600
        normal = worked_hours
        if abs(normal - 4.0) > 1e-9:
            notes.append(f"{shift}，正常時數 {fmt_hours(normal)} 小時（非 4 小時）")
    elif in_norm and not out_norm:
        shift, _ = classify_shift(in_norm)
        normal = 4.0
        notes.append(f"{shift}，無下班紀錄，計為 4 小時（default）")
    elif out_norm and not in_norm:
        shift = "早班" if (out_norm.hour < 15 or (out_norm.hour == 14 and out_norm.minute <= 30)) else "晚班" if out_norm.hour >= 20 else "未知"
        normal = 4.0
        notes.append(f"{shift}，無上班紀錄，推算計為 4 小時")

    add_record(
        records,
        summary,
        PairRecord(
            employee=name,
            date=date,
            shift=shift,
            in_raw=in_ts.strftime(TIME_FORMAT) if in_ts else "",
            in_norm=in_norm.strftime("%Y-%m-%d %H:%M") if in_norm else "",
            out_raw=out_ts.strftime(TIME_FORMAT) if out_ts else "",
            out_norm=out_norm.strftime("%Y-%m-%d %H:%M") if out_norm else "",
            normal_hours=max(normal, 0.0),
            overtime_hours=max(overtime, 0.0),
            note="；".join(notes),
        ),
        force_special=bool(notes) or inferred_no_in,
    )


def apply_daily_overtime_for_pt(records: list[PairRecord], summary: EmployeeSummary) -> None:
    """For PT employees, overtime only applies when daily total hours > 8."""
    from collections import defaultdict

    day_map: dict[str, list[PairRecord]] = defaultdict(list)
    for r in records:
        if r.employee == summary.employee:
            day_map[r.date].append(r)

    summary.normal_hours = 0.0
    summary.overtime_hours = 0.0

    for date, day_recs in sorted(day_map.items()):
        total = sum(r.normal_hours + r.overtime_hours for r in day_recs)

        if total > 8.0:
            overtime = total - 8.0
            remaining_normal = 8.0
            for r in day_recs[:-1]:
                r_total = r.normal_hours + r.overtime_hours
                r.normal_hours = min(r_total, remaining_normal)
                r.overtime_hours = 0.0
                remaining_normal = max(0.0, remaining_normal - r.normal_hours)
            day_recs[-1].normal_hours = remaining_normal
            day_recs[-1].overtime_hours = overtime
            summary.normal_hours += 8.0
            summary.overtime_hours += overtime
            summary.specials.append(
                f"{date} 日總時數 {fmt_hours(total)} 小時，計為 {fmt_hours(overtime)} 小時加班"
            )
        else:
            summary.normal_hours += total


def analyze_employee(name: str, events: list[Event], records: list[PairRecord]) -> EmployeeSummary:
    summary = EmployeeSummary(employee=name, is_full_time=name in FULL_TIME_NAMES)
    current_in: Optional[datetime] = None

    def consume_pair(in_ts: Optional[datetime], out_ts: Optional[datetime], inferred_no_in: bool = False) -> None:
        if summary.is_full_time:
            handle_full_time(summary, records, name, in_ts, out_ts, inferred_no_in)
        else:
            handle_hourly(summary, records, name, in_ts, out_ts, inferred_no_in)

    i = 0
    while i < len(events):
        e = events[i]

        if e.kind == "clock-in":
            if current_in is not None:
                if e.timestamp and abs((e.timestamp - current_in).total_seconds()) <= 60:
                    summary.specials.append(f"{current_in.date().isoformat()} 重複 clock-in（<=60秒），丟棄後者")
                    i += 1
                    continue
                consume_pair(current_in, None)
            current_in = e.timestamp

        elif e.kind == "clock-out":
            if current_in is None:
                consume_pair(None, e.timestamp, inferred_no_in=True)
            else:
                consume_pair(current_in, e.timestamp)
                current_in = None

        elif e.kind == "clock-out-no-in":
            consume_pair(None, e.timestamp, inferred_no_in=True)

        elif e.kind == "no-clock-out" and current_in is not None:
            # Handle edge case: clock-in + no-clock-out + clock-in(<=60s) + no-clock-out.
            if i + 1 < len(events):
                nxt = events[i + 1]
                if nxt.kind == "clock-in" and nxt.timestamp and abs((nxt.timestamp - current_in).total_seconds()) <= 60:
                    summary.specials.append(f"{current_in.date().isoformat()} 重複 clock-in（<=60秒），丟棄後者")
                    i += 1
            consume_pair(current_in, None)
            current_in = None

        i += 1

    if current_in is not None:
        consume_pair(current_in, None)

    if not summary.is_full_time:
        apply_daily_overtime_for_pt(records, summary)

    return summary


def extract_month_key(path: Path) -> str:
    m = re.search(r"(\d{4}-\d{2})-\d{2}~\d{4}-\d{2}-\d{2}", path.name)
    return m.group(1) if m else datetime.now().strftime("%Y-%m")


def format_summary(summaries: list[EmployeeSummary]) -> str:
    lines: list[str] = []
    for s in summaries:
        role = "正職" if s.is_full_time else "計時"
        lines.append(f"{s.employee}（{role}）:")
        lines.append(f"正常時數 {fmt_hours(s.normal_hours)} 小時")
        lines.append(f"加班時數 {fmt_hours(s.overtime_hours)} 小時")
        lines.append("特殊班別:")
        if s.specials:
            for line in s.specials:
                lines.append(f"  {line}")
        else:
            lines.append("  無")
        lines.append("")
    return "\n".join(lines)


def print_summary(summaries: list[EmployeeSummary]) -> None:
    print(format_summary(summaries))


def write_xlsx_report(records: list[PairRecord], summaries: list[EmployeeSummary], month_key: str) -> Path:
    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()

    # Sheet 1: 摘要
    ws_summary = wb.active
    ws_summary.title = "摘要"
    bold = Font(bold=True)
    row = 1
    for s in summaries:
        role = "正職" if s.is_full_time else "計時"
        cell = ws_summary.cell(row=row, column=1, value=f"{s.employee}（{role}）")
        cell.font = bold
        row += 1
        ws_summary.cell(row=row, column=1, value=f"正常時數 {fmt_hours(s.normal_hours)} 小時")
        row += 1
        ws_summary.cell(row=row, column=1, value=f"加班時數 {fmt_hours(s.overtime_hours)} 小時")
        row += 1
        ws_summary.cell(row=row, column=1, value="特殊班別:")
        row += 1
        if s.specials:
            for line in s.specials:
                ws_summary.cell(row=row, column=1, value=f"  {line}")
                row += 1
        else:
            ws_summary.cell(row=row, column=1, value="  無")
            row += 1
        row += 1  # blank row between employees

    # Sheet 2: 明細
    ws_detail = wb.create_sheet("明細")
    headers = ["員工", "班別", "日期", "上班原始", "上班normalized", "下班原始", "下班normalized", "正常時數", "加班時數", "備註"]
    for col_idx, h in enumerate(headers, 1):
        cell = ws_detail.cell(row=1, column=col_idx, value=h)
        cell.font = bold
    for row_idx, r in enumerate(records, 2):
        ws_detail.cell(row=row_idx, column=1, value=r.employee)
        ws_detail.cell(row=row_idx, column=2, value=r.shift)
        ws_detail.cell(row=row_idx, column=3, value=r.date)
        ws_detail.cell(row=row_idx, column=4, value=r.in_raw)
        ws_detail.cell(row=row_idx, column=5, value=r.in_norm)
        ws_detail.cell(row=row_idx, column=6, value=r.out_raw)
        ws_detail.cell(row=row_idx, column=7, value=r.out_norm)
        ws_detail.cell(row=row_idx, column=8, value=fmt_hours(r.normal_hours))
        ws_detail.cell(row=row_idx, column=9, value=fmt_hours(r.overtime_hours))
        ws_detail.cell(row=row_idx, column=10, value=r.note)

    # Auto-fit column widths (approximate)
    for ws in [ws_summary, ws_detail]:
        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max_len + 4, 60)

    out_dir = Path("data/clock_in_out")
    out_dir.mkdir(parents=True, exist_ok=True)
    out_path = out_dir / f"clock_report_{month_key}.xlsx"
    wb.save(out_path)
    return out_path


def write_report(records: list[PairRecord], month_key: str) -> Path:
    out_dir = Path("data/clock_in_out")
    out_dir.mkdir(parents=True, exist_ok=True)
    out_path = out_dir / f"clock_report_{month_key}.csv"
    with out_path.open("w", encoding="utf-8-sig", newline="") as f:
        w = csv.writer(f)
        w.writerow(["員工", "班別", "日期", "上班原始", "上班normalized", "下班原始", "下班normalized", "正常時數", "加班時數", "備註"])
        for r in records:
            w.writerow([r.employee, r.shift, r.date, r.in_raw, r.in_norm, r.out_raw, r.out_norm, fmt_hours(r.normal_hours), fmt_hours(r.overtime_hours), r.note])
    return out_path


def analyze_csv(csv_path: Path) -> tuple[list[PairRecord], list[EmployeeSummary], str]:
    employees = parse_csv(csv_path)
    records: list[PairRecord] = []
    summaries: list[EmployeeSummary] = []

    for name, events in employees.items():
        summary = analyze_employee(name, events, records)
        if summary.normal_hours == 0 and summary.overtime_hours == 0 and not summary.specials:
            continue
        summaries.append(summary)

    month_key = extract_month_key(csv_path)
    return records, summaries, month_key


def main() -> None:
    parser = argparse.ArgumentParser(description="Analyze iCHEF clock-in/out CSV")
    parser.add_argument("csv_path", help="Path to iCHEF clock-in/out CSV")
    args = parser.parse_args()

    records, summaries, month_key = analyze_csv(Path(args.csv_path))

    print_summary(summaries)
    out_path = write_report(records, month_key)
    print(f"CSV report generated: {out_path}")
    xlsx_path = write_xlsx_report(records, summaries, month_key)
    print(f"XLSX report generated: {xlsx_path}")


if __name__ == "__main__":
    main()
